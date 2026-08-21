#!/usr/bin/env python3
"""
Bot sprawdzający stany magazynowe (dostępność) produktów w sklepie CedarDrop.com
i wysyłający raport e-mail z PEŁNĄ listą produktów (podzieloną na kategorie)
jako załącznik CSV, plus krótkie podsumowanie zmian w treści e-maila.

WAŻNE — ZANIM URUCHOMISZ TO NA STAŁE:
CedarDrop.com to platforma B2B dropshipping (IdoSell) z tysiącami produktów.
Prawie każda polska hurtownia dropshippingowa udostępnia partnerom oficjalny
plik/feed XML lub CSV ze stanami magazynowymi (czasem też API) - to szybsze,
dokładniejsze i nie obciąża ich serwera tysiącami zapytań co 12h. Warto
zapytać obsługę CedarDrop (drop@cedardrop.com) czy taki feed jest dostępny -
to prawdopodobnie lepsze rozwiązanie niż ten skrypt.

JAK WYKRYWANA JEST DOSTĘPNOŚĆ (ważne, po realnym debugowaniu razem z użytkownikiem):
Strona produktu wysyła z serwera HTML, w którym przycisk "Dodaj do koszyka" WYGLĄDA
na aktywny niezależnie od realnego stanu magazynowego - prawdziwy stan (przycisk
zablokowany / formularz "powiadom mnie o dostępności" pokazany) jest ustawiany
DOPIERO przez JavaScript wykonywany w przeglądarce, już PO załadowaniu strony
(potwierdzone: użytkownik widział szary przycisk na żywo, a surowy HTML z tego
samego produktu pokazywał przycisk BEZ atrybutu disabled). Zwykłe pobranie strony
(requests/aiohttp) tego nie złapie - trzeba faktycznie uruchomić przeglądarkę.

Dlatego ten bot używa Playwright (headless Chromium) do sprawdzania KAŻDEGO
produktu: ładuje stronę, czeka aż JavaScript się wykona, i dopiero wtedy czyta
prawdziwy stan przycisku "Dodaj do koszyka" (atrybut disabled). To wolniejsze
i cięższe niż zwykłe pobieranie stron, ale to jedyny sposób, żeby zobaczyć to,
co realnie widzi klient. Wyszukiwanie linków do produktów w kategoriach (faza
"discovery") dalej używa szybkiego aiohttp, bo tam JS nie jest potrzebny.

Zmienne środowiskowe:
    SMTP_USER       - adres e-mail, z którego wysyłany jest raport (wymagane)
    SMTP_PASSWORD   - hasło / hasło aplikacji do tego konta (wymagane)
    EMAIL_TO        - adres, na który ma trafić raport (wymagane)
    SMTP_HOST       - domyślnie smtp.gmail.com
    SMTP_PORT       - domyślnie 587
    MAX_PRODUCTS    - opcjonalnie: ogranicz liczbę sprawdzanych produktów
                      (przydatne do testów, np. MAX_PRODUCTS=30)
    DEBUG_PRODUCT_URL - opcjonalnie: zamiast pełnego sprawdzania, otwiera JEDEN
                      podany produkt w przeglądarce i wypisuje jego prawdziwy,
                      po-JS-owy stan (przycisk, formularz powiadomień)
"""

import asyncio
import io
import json
import os
import re
import smtplib
import sys
from datetime import datetime, timezone
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import aiohttp
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright

BASE_URL = "https://cedardrop.com"
STATE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "state.json")

# Główne kategorie sklepu (nazwa, URL) - punkt startowy do znalezienia wszystkich
# produktów. Strona kategorii nadrzędnej pokazuje też produkty z podkategorii,
# więc nie trzeba osobno przechodzić każdej podkategorii.
# "Letnia wyprzedaż" jest CELOWO na końcu listy: to kategoria "przekrojowa",
# więc ma najniższy priorytet przy przypisywaniu produktu do kategorii.
# Jeśli CedarDrop doda nową kategorię główną, trzeba ją tu dopisać ręcznie.
CATEGORIES = [
    ("Zestawy prezentowe", "https://cedardrop.com/pol_m_Zestawy-prezentowe-1928.html"),
    ("Portfele męskie", "https://cedardrop.com/pol_m_Portfele-meskie-1915.html"),
    ("Portfele damskie", "https://cedardrop.com/pol_m_Portfele-damskie-1914.html"),
    ("Torebki damskie", "https://cedardrop.com/pol_m_Torebki-damskie-1916.html"),
    ("Torby i torebki męskie", "https://cedardrop.com/pol_m_Torby-i-torebki-meskie-1917.html"),
    ("Plecaki", "https://cedardrop.com/pol_m_Plecaki-1921.html"),
    ("Na laptopa", "https://cedardrop.com/pol_m_Na-laptopa-1947.html"),
    ("Paski", "https://cedardrop.com/pol_m_Paski-1933.html"),
    ("Walizki i torby podróżne", "https://cedardrop.com/pol_m_Walizki-i-torby-podrozne-1924.html"),
    ("Sakwy i akcesoria rowerowe", "https://cedardrop.com/pol_m_Sakwy-i-akcesoria-rowerowe-1934.html"),
    ("Akcesoria", "https://cedardrop.com/pol_m_Akcesoria-1925.html"),
    ("Letnia wyprzedaż", "https://cedardrop.com/pol_m_Letnia-wyprzedaz-2055.html"),
]

LINK_RE = re.compile(r'href="(/product-pol-(\d+)-[^"]+\.html)"')

CONCURRENCY = 5                 # równoległe zapytania przy SZUKANIU produktów (aiohttp, lekkie)
BROWSER_CONCURRENCY = 4         # równoległe zakładki przeglądarki przy SPRAWDZANIU (cięższe - realny Chromium)
REQUEST_TIMEOUT = 25
RETRIES = 2
DELAY_BETWEEN_REQUESTS = 0.4    # sekundy, dodatkowa uprzejmość (tylko dla fazy szukania produktów)
MAX_PAGES_PER_CATEGORY = 200    # zabezpieczenie przed nieskończoną pętlą

PAGE_TIMEOUT_MS = 25000
NETWORK_IDLE_TIMEOUT_MS = 8000  # ile najwyżej czekać, aż JS strony "się uspokoi" po załadowaniu
BLOCKED_RESOURCE_TYPES = {"image", "media", "font", "stylesheet"}  # przyspiesza ładowanie stron

HEADERS = {
    "User-Agent": "Mozilla/5.0 (compatible; CedarDropStockBot/1.0; sprawdzanie dostepnosci dla partnera dropshipping)"
}

# Selektory sprawdzane PO wykonaniu JavaScriptu (nie w surowym HTML - patrz komentarz
# na górze pliku o tym, dlaczego to konieczne). Ustalone na podstawie prawdziwego zrzutu
# strony CedarDrop dostarczonego przez użytkownika.
BUY_BUTTON_SELECTOR = "#projector_button_basket"
TELL_AVAILABILITY_SELECTOR = "#projector_tell_availability"

STATUS_LABELS = {True: "Dostępny", False: "Wyprzedany", None: "Nieznany"}


def load_state():
    if os.path.exists(STATE_FILE):
        with open(STATE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"last_checked": None, "products": {}}


def save_state(state):
    with open(STATE_FILE, "w", encoding="utf-8") as f:
        json.dump(state, f, ensure_ascii=False, indent=2)


async def fetch(session, url, semaphore):
    async with semaphore:
        last_error = None
        for attempt in range(RETRIES + 1):
            try:
                async with session.get(url, headers=HEADERS, timeout=REQUEST_TIMEOUT) as resp:
                    if resp.status == 200:
                        text = await resp.text()
                        await asyncio.sleep(DELAY_BETWEEN_REQUESTS)
                        return text
                    last_error = f"HTTP {resp.status}"
            except Exception as e:
                last_error = str(e)
            await asyncio.sleep(1.5)
        print(f"  [!] Nie udało się pobrać {url} ({last_error})")
        return None


async def discover_category(session, semaphore, category_name, cat_url, product_urls):
    counter = 0
    consecutive_misses = 0
    while consecutive_misses < 2 and counter < MAX_PAGES_PER_CATEGORY:
        page_url = cat_url if counter == 0 else f"{cat_url}?counter={counter}"
        html = await fetch(session, page_url, semaphore)
        if not html:
            consecutive_misses += 1
            counter += 1
            continue

        matches = LINK_RE.findall(html)
        new_count = 0
        for path, pid in matches:
            if pid not in product_urls:
                product_urls[pid] = {"url": BASE_URL + path, "category": category_name}
                new_count += 1

        if not matches:
            consecutive_misses += 1
        else:
            consecutive_misses = 0

        counter += 1

    print(f"  {category_name}: suma dotychczasowa produktów = {len(product_urls)}")


async def discover_products(session, semaphore):
    product_urls = {}
    for category_name, cat_url in CATEGORIES:
        await discover_category(session, semaphore, category_name, cat_url, product_urls)
    return product_urls


def sample_across_categories(product_urls, limit):
    """Do testów (MAX_PRODUCTS): wybiera próbkę rozłożoną po kategoriach (round-robin),
    zamiast pierwszych z rzędu (które przy sporym MAX_PRODUCTS i tak wypadałyby
    z jednej, pierwszej przetworzonej kategorii - słaby test różnych szablonów stron)."""
    by_category = {}
    for pid, info in product_urls.items():
        by_category.setdefault(info["category"], []).append((pid, info))
    lists = list(by_category.values())

    sampled = {}
    idx = 0
    while len(sampled) < limit and idx < max((len(l) for l in lists), default=0):
        for lst in lists:
            if idx < len(lst):
                pid, info = lst[idx]
                sampled[pid] = info
                if len(sampled) >= limit:
                    break
        idx += 1
    return sampled


async def _new_page(browser):
    """Nowa zakładka z zablokowanymi obrazkami/czcionkami/CSS - dużo szybsze
    ładowanie, a nam potrzebny jest tylko HTML+JS, nie wygląd strony."""
    page = await browser.new_page(user_agent=HEADERS["User-Agent"])

    async def _block(route):
        if route.request.resource_type in BLOCKED_RESOURCE_TYPES:
            await route.abort()
        else:
            await route.continue_()

    await page.route("**/*", _block)
    return page


async def check_product(browser, url, pid, category):
    """Otwiera stronę produktu w prawdziwej przeglądarce, czeka aż JavaScript
    się wykona, i czyta REALNY stan przycisku 'Dodaj do koszyka' - patrz duży
    komentarz na górze pliku o tym, dlaczego to konieczne."""
    last_error = None
    for attempt in range(RETRIES + 1):
        page = None
        try:
            page = await _new_page(browser)
            await page.goto(url, wait_until="domcontentloaded", timeout=PAGE_TIMEOUT_MS)
            try:
                await page.wait_for_load_state("networkidle", timeout=NETWORK_IDLE_TIMEOUT_MS)
            except Exception:
                pass  # niekrytyczne - jedziemy dalej z tym co już się załadowało

            h1 = page.locator("h1").first
            name = (await h1.text_content()) if await h1.count() else None
            name = name.strip() if name else pid

            in_stock = None
            btn = page.locator(BUY_BUTTON_SELECTOR)
            if await btn.count() > 0:
                in_stock = not await btn.first.is_disabled()
            else:
                # zapasowo: brak przycisku - sprawdź czy widoczny jest formularz "powiadom mnie"
                tell = page.locator(TELL_AVAILABILITY_SELECTOR)
                if await tell.count() > 0:
                    in_stock = not await tell.first.is_visible()

            return {"id": pid, "name": name, "url": url, "category": category, "in_stock": in_stock}
        except Exception as e:
            last_error = str(e)
        finally:
            if page:
                await page.close()
    print(f"  [!] Nie udało się sprawdzić {url} ({last_error})")
    return {"id": pid, "name": pid, "url": url, "category": category, "in_stock": None, "error": True}


async def debug_product(url):
    """Tryb debugowania (patrz DEBUG_PRODUCT_URL w main()): otwiera JEDEN
    podany produkt w prawdziwej przeglądarce i wypisuje jego rzeczywisty,
    PO-JS-owy stan (to, co faktycznie widziałby klient), żeby dało się
    sprawdzić, czy wykrywanie działa poprawnie na konkretnym przykładzie.
    Kończy działanie od razu po wypisaniu - nie odpala pełnego sprawdzania."""
    async with async_playwright() as p:
        browser = await p.chromium.launch()
        try:
            page = await _new_page(browser)
            resp = await page.goto(url, wait_until="domcontentloaded", timeout=PAGE_TIMEOUT_MS)
            print(f"HTTP status: {resp.status if resp else '???'}")
            try:
                await page.wait_for_load_state("networkidle", timeout=NETWORK_IDLE_TIMEOUT_MS)
            except Exception:
                print("(uwaga: przekroczony czas oczekiwania na 'uspokojenie się' sieci - wynik mimo to poniżej)")

            h1 = page.locator("h1").first
            name = (await h1.text_content()) if await h1.count() else "???"
            print(f"Nazwa produktu: {name.strip() if name else '???'}\n")

            btn = page.locator(BUY_BUTTON_SELECTOR)
            if await btn.count() > 0:
                disabled = await btn.first.is_disabled()
                btn_text = (await btn.first.text_content() or "").strip()
                print(f"Przycisk '{BUY_BUTTON_SELECTOR}': znaleziony, tekst='{btn_text}', disabled={disabled}")
                print(f"  -> wg tego: {'WYPRZEDANY' if disabled else 'DOSTĘPNY'}")
            else:
                print(f"Przycisk '{BUY_BUTTON_SELECTOR}': NIE ZNALEZIONY na stronie")

            tell = page.locator(TELL_AVAILABILITY_SELECTOR)
            if await tell.count() > 0:
                visible = await tell.first.is_visible()
                print(f"Formularz '{TELL_AVAILABILITY_SELECTOR}': znaleziony, widoczny={visible}")
            else:
                print(f"Formularz '{TELL_AVAILABILITY_SELECTOR}': NIE ZNALEZIONY na stronie")
        except Exception as e:
            print(f"BŁĄD: {e}")
        finally:
            await browser.close()


async def check_all_products(browser, product_urls):
    results = {}
    semaphore = asyncio.Semaphore(BROWSER_CONCURRENCY)

    async def worker(pid, info):
        async with semaphore:
            results[pid] = await check_product(browser, info["url"], pid, info["category"])

    tasks = [asyncio.create_task(worker(pid, info)) for pid, info in product_urls.items()]
    total = len(tasks)
    done = 0
    for coro in asyncio.as_completed(tasks):
        await coro
        done += 1
        if done % 50 == 0 or done == total:
            print(f"  Sprawdzono {done}/{total} produktów...")
    return results


def compute_changes(old_products, new_results):
    """pid -> etykieta zmiany: '', 'nowy w ofercie', 'wypadł z magazynu', 'wrócił do magazynu'.
    Zmianę zgłaszamy tylko, gdy oba odczyty (stary i nowy) są jednoznaczne (True/False) -
    status "Nieznany" nigdy nie generuje fałszywej zmiany."""
    changes = {}
    for pid, new in new_results.items():
        if new.get("error"):
            continue
        old = old_products.get(pid)
        if old is None:
            changes[pid] = "nowy w ofercie"
            continue
        old_stock, new_stock = old.get("in_stock"), new.get("in_stock")
        if old_stock is True and new_stock is False:
            changes[pid] = "wypadł z magazynu"
        elif old_stock is False and new_stock is True:
            changes[pid] = "wrócił do magazynu"
        else:
            changes[pid] = ""
    return changes


SHEET_HEADERS = ["Nazwa produktu", "Status", "Zmiana od ostatniego raportu", "Link do produktu"]
SHEET_COL_WIDTHS = [50, 14, 26, 60]

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4A5568", end_color="4A5568", fill_type="solid")
NORMAL_FONT = Font(name="Arial", size=10)
BOLD_FONT = Font(name="Arial", size=10, bold=True)
LINK_FONT = Font(name="Arial", size=10, color="1155CC", underline="single")
STATUS_FONTS = {
    "Dostępny": Font(name="Arial", size=10, color="1B7A3D", bold=True),
    "Wyprzedany": Font(name="Arial", size=10, color="C0392B", bold=True),
    "Nieznany": Font(name="Arial", size=10, color="8A6D00", bold=True),
    "Błąd sprawdzania": Font(name="Arial", size=10, color="8A6D00"),
}


def _style_header_row(ws, n_cols):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"


def _write_category_sheet(wb, title, rows):
    """rows: lista (pid, wynik) dla jednej kategorii, wynik jak z new_results + 'change'."""
    ws = wb.create_sheet(title=title[:31])
    ws.append(SHEET_HEADERS)
    _style_header_row(ws, len(SHEET_HEADERS))

    for pid, r, change in sorted(rows, key=lambda t: t[1].get("name", "")):
        status = "Błąd sprawdzania" if r.get("error") else STATUS_LABELS[r["in_stock"]]
        row = [r.get("name", pid), status, change, r.get("url", "")]
        ws.append(row)
        row_idx = ws.max_row
        for col in (1, 3):
            ws.cell(row=row_idx, column=col).font = NORMAL_FONT
        ws.cell(row=row_idx, column=2).font = STATUS_FONTS.get(status, NORMAL_FONT)
        link_cell = ws.cell(row=row_idx, column=4)
        if r.get("url"):
            link_cell.hyperlink = r["url"]
        link_cell.font = LINK_FONT

    for i, width in enumerate(SHEET_COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    if ws.max_row >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(SHEET_HEADERS))}{ws.max_row}"
    return ws


def _write_summary_sheet(wb, new_results, changes, errors_count, is_first_run, total_tracked, now_label):
    ws = wb.create_sheet(title="Podsumowanie")
    ws["A1"] = "Raport stanów magazynowych CedarDrop.com"
    ws["A1"].font = Font(name="Arial", size=14, bold=True)
    ws["A2"] = now_label
    ws["A2"].font = Font(name="Arial", size=10, italic=True, color="666666")

    per_cat = {}
    for r in new_results.values():
        if r.get("error"):
            continue
        cat = r.get("category") or "(brak kategorii)"
        d = per_cat.setdefault(cat, {"in": 0, "out": 0, "unknown": 0})
        if r["in_stock"] is True:
            d["in"] += 1
        elif r["in_stock"] is False:
            d["out"] += 1
        else:
            d["unknown"] += 1

    headers = ["Kategoria (zakładka)", "Dostępne", "Niedostępne"]
    header_row = 4
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    ws.freeze_panes = f"A{header_row + 1}"

    row_idx = header_row + 1
    total_in = total_out = total_unknown = 0
    for cat_name, _ in CATEGORIES:
        d = per_cat.get(cat_name, {"in": 0, "out": 0, "unknown": 0})
        total_in += d["in"]; total_out += d["out"]; total_unknown += d["unknown"]
        values = [cat_name, d["in"], d["out"]]
        for col, v in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col, value=v).font = NORMAL_FONT
        row_idx += 1

    total_row = row_idx
    totals = ["RAZEM", total_in, total_out]
    for col, v in enumerate(totals, start=1):
        ws.cell(row=total_row, column=col, value=v).font = BOLD_FONT
    ws.auto_filter.ref = f"A{header_row}:C{total_row}"

    row_idx = total_row + 2
    if is_first_run:
        ws.cell(row=row_idx, column=1, value="To pierwsze sprawdzenie - brak jeszcze danych o zmianach.").font = NORMAL_FONT
    else:
        newly_out = sum(1 for c in changes.values() if c == "wypadł z magazynu")
        newly_in = sum(1 for c in changes.values() if c == "wrócił do magazynu")
        new_products = sum(1 for c in changes.values() if c == "nowy w ofercie")
        ws.cell(row=row_idx, column=1, value="Zmiany od ostatniego sprawdzenia:").font = BOLD_FONT
        row_idx += 1
        for label, val in [("Wypadło z magazynu", newly_out), ("Wróciło do magazynu", newly_in), ("Nowe w ofercie", new_products)]:
            ws.cell(row=row_idx, column=1, value=label).font = NORMAL_FONT
            ws.cell(row=row_idx, column=2, value=val).font = NORMAL_FONT
            row_idx += 1

    if errors_count:
        row_idx += 1
        ws.cell(row=row_idx, column=1, value=f"Nie udało się sprawdzić {errors_count} produktów (błąd sieci) - spróbuję ponownie następnym razem.").font = NORMAL_FONT

    for i, width in enumerate([28, 12, 14], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    return ws


def build_xlsx(new_results, changes, errors_count, is_first_run, total_tracked):
    """Jeden plik .xlsx: zakładka 'Podsumowanie' + jedna zakładka na kategorię
    (w kolejności z CATEGORIES), żeby dało się kliknąć np. od razu w 'Akcesoria'
    albo 'Letnia wyprzedaż' zamiast przewijać jedną wspólną listę."""
    now_label = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

    by_category = {}
    for pid, r in new_results.items():
        cat = r.get("category") or "(brak kategorii)"
        by_category.setdefault(cat, []).append((pid, r, changes.get(pid, "")))

    wb = Workbook()
    wb.remove(wb.active)  # domyślny pusty arkusz - zastąpimy własnymi, w naszej kolejności

    _write_summary_sheet(wb, new_results, changes, errors_count, is_first_run, total_tracked, now_label)
    for cat_name, _ in CATEGORIES:
        _write_category_sheet(wb, cat_name, by_category.get(cat_name, []))

    wb.active = 0  # przy otwarciu pliku od razu widoczne Podsumowanie
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_summary_body(new_results, changes, errors_count, is_first_run, total_tracked, xlsx_filename):
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    lines = [f"Raport stanów magazynowych CedarDrop.com — {now}", ""]

    per_cat = {}
    for r in new_results.values():
        if r.get("error"):
            continue
        cat = r.get("category") or "(brak kategorii)"
        d = per_cat.setdefault(cat, {"in": 0, "out": 0, "unknown": 0})
        if r["in_stock"] is True:
            d["in"] += 1
        elif r["in_stock"] is False:
            d["out"] += 1
        else:
            d["unknown"] += 1

    lines.append("Stan wg kategorii (dostępne / niedostępne / nieznane):")
    for cat in sorted(per_cat):
        d = per_cat[cat]
        total = d["in"] + d["out"] + d["unknown"]
        extra = f" / {d['unknown']} nieznanych" if d["unknown"] else ""
        lines.append(f"  {cat}: {d['in']} / {d['out']}{extra}  (razem {total})")
    lines.append("")

    newly_out = sum(1 for c in changes.values() if c == "wypadł z magazynu")
    newly_in = sum(1 for c in changes.values() if c == "wrócił do magazynu")
    new_products = sum(1 for c in changes.values() if c == "nowy w ofercie")

    if is_first_run:
        lines.append("To pierwsze sprawdzenie - pełna lista wszystkich produktów jest w załączonym pliku Excel (jedna zakładka na kategorię).")
    else:
        lines.append(
            f"Zmiany od ostatniego sprawdzenia: {newly_out} wypadło z magazynu, "
            f"{newly_in} wróciło, {new_products} nowych w ofercie."
        )
        lines.append("Pełna lista WSZYSTKICH produktów (z kolumną \"Zmiana\") jest w załączonym pliku Excel - osobna zakładka na każdą kategorię, plus zakładka Podsumowanie na start.")

    if errors_count:
        lines.append(f"Nie udało się sprawdzić {errors_count} produktów (błąd sieci) - spróbuję ponownie następnym razem.")

    lines.append("")
    lines.append(f"Łącznie śledzonych produktów: {total_tracked}")
    lines.append(f"Załącznik: {xlsx_filename}")

    return "\n".join(lines)


def send_email(subject, body, attachment_bytes=None, attachment_filename=None):
    smtp_host = os.environ.get("SMTP_HOST") or "smtp.gmail.com"
    smtp_port = int(os.environ.get("SMTP_PORT") or "587")
    smtp_user = os.environ["SMTP_USER"]
    smtp_password = os.environ["SMTP_PASSWORD"]
    email_to = os.environ["EMAIL_TO"]

    msg = MIMEMultipart()
    msg["From"] = smtp_user
    msg["To"] = email_to
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain", "utf-8"))

    if attachment_bytes is not None:
        part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        part.set_payload(attachment_bytes)
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f'attachment; filename="{attachment_filename}"')
        msg.attach(part)

    with smtplib.SMTP(smtp_host, smtp_port) as server:
        server.starttls()
        server.login(smtp_user, smtp_password)
        server.sendmail(smtp_user, [email_to], msg.as_string())


def check_required_env():
    missing = [k for k in ("SMTP_USER", "SMTP_PASSWORD", "EMAIL_TO") if not os.environ.get(k)]
    if missing:
        print(f"BŁĄD: brakuje zmiennych środowiskowych: {', '.join(missing)}")
        print("Ustaw je jako sekrety repozytorium (Settings -> Secrets and variables -> Actions).")
        sys.exit(1)


async def main():
    check_required_env()
    print(f"=== Start sprawdzania: {datetime.now(timezone.utc).isoformat()} ===")

    state = load_state()
    is_first_run = state.get("last_checked") is None

    semaphore = asyncio.Semaphore(CONCURRENCY)
    async with aiohttp.ClientSession() as session:
        print("Krok 1/3: szukam produktów w kategoriach...")
        product_urls = await discover_products(session, semaphore)
        print(f"Znaleziono {len(product_urls)} unikalnych produktów.")

    max_products = os.environ.get("MAX_PRODUCTS")
    if max_products:
        limit = int(max_products)
        product_urls = sample_across_categories(product_urls, limit)
        print(f"UWAGA: tryb testowy (MAX_PRODUCTS={limit}), próbka rozłożona po kategoriach")

    print("Krok 2/3: sprawdzam dostępność każdego produktu w przeglądarce (to potrwa dłużej niż szukanie linków)...")
    async with async_playwright() as p:
        browser = await p.chromium.launch()
        try:
            new_results = await check_all_products(browser, product_urls)
        finally:
            await browser.close()

    print("Krok 3/3: buduję CSV, porównuję ze stanem poprzednim i wysyłam raport...")
    changes = compute_changes(state.get("products", {}), new_results)
    errors_count = sum(1 for r in new_results.values() if r.get("error"))
    unknown_count = sum(1 for r in new_results.values() if not r.get("error") and r.get("in_stock") is None)
    if unknown_count:
        print(f"  [i] {unknown_count} produktów bez jednoznacznego wskaźnika dostępności (status: Nieznany)")

    now_str = datetime.now(timezone.utc).strftime("%Y-%m-%d_%H%M")
    xlsx_filename = f"cedardrop_stany_{now_str}.xlsx"
    xlsx_bytes = build_xlsx(new_results, changes, errors_count, is_first_run, len(product_urls))

    out_count = sum(1 for r in new_results.values() if r.get("in_stock") is False)
    total = len(product_urls)
    if is_first_run:
        subject = f"Raport CedarDrop: pierwsze sprawdzenie ({total} produktów)"
    else:
        newly_out = sum(1 for c in changes.values() if c == "wypadł z magazynu")
        newly_in = sum(1 for c in changes.values() if c == "wrócił do magazynu")
        subject = f"Raport CedarDrop: {out_count}/{total} niedostępnych ({newly_out} nowo, {newly_in} wróciło)"

    body = build_summary_body(new_results, changes, errors_count, is_first_run, total, xlsx_filename)

    try:
        send_email(subject, body, attachment_bytes=xlsx_bytes, attachment_filename=xlsx_filename)
        print("E-mail z załącznikiem XLSX wysłany.")
    except Exception as e:
        print(f"BŁĄD wysyłki e-maila: {e}")

    # Zachowaj poprzedni znany stan dla produktów, których nie udało się
    # sprawdzić w tym przebiegu (błąd sieci), żeby ich nie "zgubić".
    merged_products = dict(state.get("products", {}))
    for pid, r in new_results.items():
        if r.get("error"):
            continue
        merged_products[pid] = {
            "name": r["name"], "url": r["url"], "category": r.get("category", ""),
            "in_stock": r["in_stock"],
        }

    save_state({
        "last_checked": datetime.now(timezone.utc).isoformat(),
        "products": merged_products,
    })
    print("=== Koniec ===")


if __name__ == "__main__":
    debug_url = os.environ.get("DEBUG_PRODUCT_URL")
    if debug_url:
        asyncio.run(debug_product(debug_url))
    else:
        asyncio.run(main())
